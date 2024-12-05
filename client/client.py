import argparse
import requests
import pandas as pd
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# Renkleri tanımlama
green = PatternFill(start_color="007500", end_color="007500", fill_type="solid")
orange = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")
red = PatternFill(start_color="b30000", end_color="b30000", fill_type="solid")


def color_by_hu(hu_date):
    """Muayene tarihine göre renk döndür."""
    if pd.isna(hu_date):
        return None
    diff_months = (pd.Timestamp.now() - pd.to_datetime(hu_date)).days // 30
    if diff_months <= 3:
        return green
    elif diff_months <= 12:
        return orange
    else:
        return red


def main():
    # Komut satırı argümanlarını al
    parser = argparse.ArgumentParser(description="CSV'yi REST API'ye gönder ve Excel'e dönüştür")
    parser.add_argument('-f', '--file', required=True, help="Gönderilecek CSV dosyası")
    parser.add_argument('-k', '--keys', nargs='+', required=True, help="Ekstra sütunlar")
    parser.add_argument('-c', '--colored', action='store_true', default=True, help="Satırları renklendir")
    args = parser.parse_args()

    # 1. CSV'yi sunucuya gönder
    with open(args.file, 'rb') as f:
        response = requests.post('http://127.0.0.1:8000/api/upload/', files={'file': f})
        if response.status_code != 200:
            print(f"Hata: Sunucu {response.status_code} döndürdü.")
            return

        data = response.json()

    # 2. Sunucudan gelen verileri DataFrame'e dönüştür
    df = pd.DataFrame(data)

    # 3. Gerekli sütunları seç ve sıralama yap
    required_columns = ['rnr', 'gruppe'] + args.keys
    df = df[required_columns]
    df = df.sort_values(by='gruppe')

    # 4. Excel dosyasını oluştur
    filename = f"vehicles_{datetime.now().strftime('%Y-%m-%d')}.xlsx"
    df.to_excel(filename, index=False)

    # 5. Renkli çıktı işleme
    if args.colored:
        wb = load_workbook(filename)
        sheet = wb.active

        for idx, row in df.iterrows():
            # Hu sütunundan renklendirme uygula
            fill_color = color_by_hu(row.get('hu'))
            if fill_color:
                for cell in sheet[idx + 2]:  # Excel satırları 1 tabanlıdır, başlığı atlamak için +2
                    cell.fill = fill_color

        wb.save(filename)

    print(f"Excel dosyası başarıyla oluşturuldu: {filename}")


if __name__ == '__main__':
    main()
